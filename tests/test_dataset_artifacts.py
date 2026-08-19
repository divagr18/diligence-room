"""Planted-fact tests for the committed synthetic dataset artifacts (D2-M4/S5).

The contract PDF and FY27 financials carry the keystone demo facts; these
tests pin them byte-exactly so later gates (evidence gate, golden set) can
quote them verbatim.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from scripts.author_dataset import (
    MERIDIAN_REVENUE,
    TOTAL_REVENUE,
    write_amendment_2030,
    write_contract_meridian_logistics,
    write_financials_fy27,
    write_hr_roster,
    write_scanned_invoice,
    write_tech_inventory,
    write_vendor_agreement_2027,
)

DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "vantage_robotics"
SCENARIOS_DIR = Path(__file__).resolve().parent.parent / "data" / "scenarios"
CONTRACT_PDF = DATA_DIR / "contract_meridian_logistics.pdf"
FINANCIALS_XLSX = DATA_DIR / "financials_fy27.xlsx"
HR_ROSTER_XLSX = DATA_DIR / "hr_roster_vantage.xlsx"
TECH_INVENTORY_PDF = DATA_DIR / "tech_inventory.pdf"
VENDOR_AGREEMENT_PDF = DATA_DIR / "vendor_agreement_2027.pdf"
SCANNED_INVOICE_PDF = SCENARIOS_DIR / "scanned_invoice.pdf"

ROSTER_DATE = date(2026, 8, 14)
WHITFIELD_DEPARTURE = date(2026, 10, 13)

COC_SPAN = (
    "may terminate this Agreement by written notice delivered within ninety "
    "(90) days following a Change of Control"
)

EXPECTED_REVENUES: dict[str, int] = {
    "Meridian Logistics, Inc.": 8_893_800,
    "Halbrook Manufacturing": 12_400_000,
    "Cascade Retail Group": 9_850_000,
    "Public Sector & Municipal": 8_106_200,
    "Other (43 accounts)": 9_350_000,
}
EXPECTED_TOTAL = 48_600_000


def _normalized_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    raw = "\n".join(page.extract_text() or "" for page in reader.pages)
    return " ".join(raw.split())


class TestContractArtifact:
    def test_coc_span_extractable_byte_exact(self) -> None:
        assert COC_SPAN in _normalized_pdf_text(CONTRACT_PDF)

    def test_meridian_is_a_party(self) -> None:
        assert "Meridian Logistics" in _normalized_pdf_text(CONTRACT_PDF)

    def test_clause_locator_present(self) -> None:
        assert "11.3" in _normalized_pdf_text(CONTRACT_PDF)

    def test_regeneration_is_byte_identical(self, tmp_path: Path) -> None:
        regenerated = tmp_path / "contract_meridian_logistics.pdf"
        write_contract_meridian_logistics(regenerated)
        assert regenerated.read_bytes() == CONTRACT_PDF.read_bytes()


class TestFinancialsArtifact:
    def _revenue_rows(self, path: Path) -> dict[str, int]:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook["FY27 Projected Revenue"]
        rows: dict[str, int] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            customer, _segment, revenue, _percent = row[0], row[1], row[2], row[3]
            if customer is None or customer == "TOTAL":
                continue
            assert isinstance(revenue, (int, float))
            rows[str(customer)] = int(revenue)
        return rows

    def _total_row(self, path: Path) -> int:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook["FY27 Projected Revenue"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == "TOTAL":
                total = row[2]
                assert isinstance(total, (int, float))
                return int(total)
        raise AssertionError("TOTAL row missing from FY27 Projected Revenue sheet")

    def test_exact_customer_revenues(self) -> None:
        assert self._revenue_rows(FINANCIALS_XLSX) == EXPECTED_REVENUES

    def test_total_and_share_of_customer_x(self) -> None:
        total = self._total_row(FINANCIALS_XLSX)
        meridian = self._revenue_rows(FINANCIALS_XLSX)["Meridian Logistics, Inc."]
        assert total == EXPECTED_TOTAL
        assert sum(EXPECTED_REVENUES.values()) == EXPECTED_TOTAL
        ratio = meridian / total
        assert ratio == 0.183
        assert f"{ratio:.3%}" == "18.300%"

    def test_regeneration_preserves_values(self, tmp_path: Path) -> None:
        regenerated = tmp_path / "financials_fy27.xlsx"
        write_financials_fy27(regenerated)
        assert self._revenue_rows(regenerated) == self._revenue_rows(FINANCIALS_XLSX)
        assert self._total_row(regenerated) == self._total_row(FINANCIALS_XLSX)


class TestHrRosterAndTechInventory:
    def _whitfield_row(self, path: Path) -> tuple[object, ...]:
        workbook = load_workbook(path)
        sheet = workbook["Roster"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == "Dana Whitfield":
                return row
        raise AssertionError("Dana Whitfield missing from HR roster")

    def test_hr_roster_finalized_without_draft_flag(self) -> None:
        workbook = load_workbook(HR_ROSTER_XLSX)
        assert "Roster" in workbook.sheetnames
        assert not any("DRAFT" in name for name in workbook.sheetnames)

    def test_hr_roster_contains_whitfield_resignation(self) -> None:
        whitfield = self._whitfield_row(HR_ROSTER_XLSX)
        assert whitfield[1] == "VP Customer Success"
        assert whitfield[3] == "Meridian Logistics, Inc."
        assert whitfield[4] == "Resigning"
        departure = whitfield[5]
        assert isinstance(departure, datetime)
        assert departure.date() == WHITFIELD_DEPARTURE
        assert (departure.date() - ROSTER_DATE).days == 60

    def test_hr_roster_regeneration_preserves_whitfield(self, tmp_path: Path) -> None:
        regenerated = tmp_path / "hr_roster_vantage.xlsx"
        write_hr_roster(regenerated)
        assert self._whitfield_row(regenerated) == self._whitfield_row(HR_ROSTER_XLSX)

    def test_tech_inventory_final_status_present_and_draft_gone(self) -> None:
        text = _normalized_pdf_text(TECH_INVENTORY_PDF)
        assert "FINAL" in text
        assert "DRAFT" not in text
        assert "TitanBridge 4.1" in text
        assert "Meridian" in text

    def test_tech_inventory_regeneration_byte_identical(self, tmp_path: Path) -> None:
        regenerated = tmp_path / "tech_inventory.pdf"
        write_tech_inventory(regenerated)
        assert regenerated.read_bytes() == TECH_INVENTORY_PDF.read_bytes()


class TestTechInventoryFactStability:
    """The TitanBridge fact must survive DRAFT->FINAL byte-identically."""

    TITANBRIDGE_FACTS = (
        (
            "Runs on TitanBridge 4.1 (vendor end-of-life 2026-03; no support "
            "contract). Serves the Meridian Logistics, Inc. account. Migration "
            "to a supported runtime is estimated at 9-12 months."
        ),
        (
            "Proprietary computer-vision pipeline; patents assigned to Vantage "
            "Robotics, Inc. Open-source components under Apache-2.0."
        ),
        "Firmware maintained in-house; hardware refresh due FY28.",
    )

    def test_entry_texts_present_unchanged(self) -> None:
        text = _normalized_pdf_text(TECH_INVENTORY_PDF)
        for fact in self.TITANBRIDGE_FACTS:
            assert " ".join(fact.split()) in text


class TestVendorAgreement2027:
    def test_titanbridge_license_present(self) -> None:
        text = _normalized_pdf_text(VENDOR_AGREEMENT_PDF)
        assert "TitanBridge" in text
        assert "Vantage Robotics" in text
        assert "TitanBridge Systems" in text

    def test_exclusivity_ends_2027_06_30(self) -> None:
        text = _normalized_pdf_text(VENDOR_AGREEMENT_PDF)
        assert "2027-06-30" in text

    def test_clause_numbering_present(self) -> None:
        text = _normalized_pdf_text(VENDOR_AGREEMENT_PDF)
        assert "4. Exclusivity" in text

    def test_eol_acknowledged(self) -> None:
        text = _normalized_pdf_text(VENDOR_AGREEMENT_PDF)
        assert "end-of-life" in text

    def test_regeneration_byte_identical(self, tmp_path: Path) -> None:
        regenerated = tmp_path / "vendor_agreement_2027.pdf"
        write_vendor_agreement_2027(regenerated)
        assert regenerated.read_bytes() == VENDOR_AGREEMENT_PDF.read_bytes()


class TestScannedInvoice:
    def test_image_only_pdf_has_no_extractable_text(self) -> None:
        reader = PdfReader(str(SCANNED_INVOICE_PDF))
        assert reader.pages
        for page in reader.pages:
            assert not (page.extract_text() or "").strip()

    def test_regeneration_byte_identical(self, tmp_path: Path) -> None:
        regenerated = tmp_path / "scanned_invoice.pdf"
        write_scanned_invoice(regenerated)
        assert regenerated.read_bytes() == SCANNED_INVOICE_PDF.read_bytes()


class TestRegenerationDeterminism:
    def test_financials_regenerate_byte_identical(self, tmp_path: Path) -> None:
        regenerated = tmp_path / "financials_fy27.xlsx"
        write_financials_fy27(regenerated)
        assert regenerated.read_bytes() == FINANCIALS_XLSX.read_bytes()

    def test_hr_roster_regenerates_byte_identical(self, tmp_path: Path) -> None:
        regenerated = tmp_path / "hr_roster_vantage.xlsx"
        write_hr_roster(regenerated)
        assert regenerated.read_bytes() == HR_ROSTER_XLSX.read_bytes()


AMENDMENT_PDF = DATA_DIR / "amendment_2030.pdf"


class TestAmendment2030:
    def test_exclusivity_extended_to_2030(self) -> None:
        text = _normalized_pdf_text(AMENDMENT_PDF)
        assert "2030-06-30" in text

    def test_amends_section_4_only(self) -> None:
        text = _normalized_pdf_text(AMENDMENT_PDF)
        assert "Section 4" in text
        assert "Exclusivity" in text
        assert "full force and effect" in text

    def test_references_original_agreement(self) -> None:
        text = _normalized_pdf_text(AMENDMENT_PDF)
        assert "SOFTWARE LICENSE AGREEMENT" in text
        assert "TitanBridge" in text

    def test_original_vendor_agreement_unchanged(self) -> None:
        vendor_text = _normalized_pdf_text(VENDOR_AGREEMENT_PDF)
        assert "2027-06-30" in vendor_text

    def test_regeneration_byte_identical(self, tmp_path: Path) -> None:
        regenerated = tmp_path / "amendment_2030.pdf"
        write_amendment_2030(regenerated)
        assert regenerated.read_bytes() == AMENDMENT_PDF.read_bytes()


SCAFFOLD_DOCS: dict[str, tuple[str, ...]] = {
    "tax_exposure.pdf": ("tax exposure", "carryforward", "open tax years"),
    "regulatory_correspondence.pdf": ("regulatory", "permit", "market concentration"),
    "esg_report.pdf": ("emissions", "disclosure", "environmental liability"),
    "lease_meridian.pdf": ("lease", "renewal", "Meridian"),
}


class TestScaffoldDocs:
    def test_all_four_scaffold_docs_exist_with_markers(self) -> None:
        for name, markers in SCAFFOLD_DOCS.items():
            text = _normalized_pdf_text(DATA_DIR / name)
            for marker in markers:
                assert marker in text, f"{name} missing marker {marker!r}"

    def test_lease_doc_has_coc_provision(self) -> None:
        text = _normalized_pdf_text(DATA_DIR / "lease_meridian.pdf")
        assert "change of control" in text.lower()

    def test_scaffold_docs_regenerate_byte_identical(self, tmp_path: Path) -> None:
        from scripts.author_dataset import (
            write_esg_report,
            write_lease_meridian,
            write_regulatory_correspondence,
            write_tax_exposure,
        )

        writers = {
            "tax_exposure.pdf": write_tax_exposure,
            "regulatory_correspondence.pdf": write_regulatory_correspondence,
            "esg_report.pdf": write_esg_report,
            "lease_meridian.pdf": write_lease_meridian,
        }
        for name, writer in writers.items():
            regenerated = tmp_path / name
            writer(regenerated)
            assert regenerated.read_bytes() == (DATA_DIR / name).read_bytes(), (
                f"{name} regeneration not byte-identical"
            )


class TestDatasetScriptCli:
    """The generator is a write tool: unknown arguments must refuse, not regenerate."""

    def test_unknown_argument_refuses(self) -> None:
        import subprocess
        import sys

        result = subprocess.run(
            [sys.executable, "scripts/author_dataset.py", "--bogus-flag"],
            capture_output=True,
            text=True,
            timeout=300,
        )
        assert result.returncode != 0

    def test_help_exits_zero_without_arguments(self) -> None:
        import subprocess
        import sys

        result = subprocess.run(
            [sys.executable, "scripts/author_dataset.py", "--help"],
            capture_output=True,
            text=True,
            timeout=300,
        )
        assert result.returncode == 0
        assert "usage" in result.stdout.lower()


class TestDatasetConsistency:
    def test_customer_x_share_exact_in_workbook(self) -> None:
        financials = TestFinancialsArtifact()
        meridian = financials._revenue_rows(FINANCIALS_XLSX)["Meridian Logistics, Inc."]
        total = financials._total_row(FINANCIALS_XLSX)
        assert meridian == MERIDIAN_REVENUE
        assert total == TOTAL_REVENUE
        assert meridian / total == 0.183
        assert f"{meridian / total:.3%}" == "18.300%"

    def test_plan_doc_pins_the_same_figures(self) -> None:
        plan = (DATA_DIR / "DATASET_PLAN.md").read_text(encoding="utf-8")
        assert "8,893,800" in plan
        assert "48,600,000" in plan
        assert "18.300%" in plan
