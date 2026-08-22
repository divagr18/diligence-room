"""Golden-set validation tests (BUILD_PLAN D12-M1).

Pins the golden set against the committed clean corpus: twenty unique docs
that exactly cover the parseable dataset artifacts, every doc_id present on
disk and parseable (or honestly flagged ``needs_ocr``), the keystone planted
facts byte-exact (CoC span at clause:11.3, 18.3% concentration, Whitfield
+60d, TitanBridge EOL locator), zero lorem filler, and keystone expectations
identical to the offline fleet producers.
"""

from __future__ import annotations

from datetime import datetime

import pytest
from openpyxl import load_workbook

from agents.fleet import DEEP_WORKSTREAM_DOCUMENTS, extract_fact
from evals.golden_set import (
    COC_CLAUSE_LOCATOR,
    COC_SPAN,
    DATA_DIR,
    FY27_REVENUE_SHEET,
    GOLDEN_CONCENTRATION_RATIO,
    GOLDEN_SET,
    ROSTER_REFERENCE_DATE,
    SCENARIOS_DIR,
    WHITFIELD_DEPARTURE,
    GoldenDoc,
    golden_doc,
    golden_path,
    load_golden_set,
)
from ingestion.chunking import chunk
from ingestion.parsing import LocalParser
from registry.models import Workstream
from scripts.author_dataset import MERIDIAN_REVENUE, TOTAL_REVENUE

DEAL = "deal-falcon"
_PARSEABLE_SUFFIXES = frozenset({".pdf", ".xlsx", ".docx", ".eml"})
_REDTEAM_FIXTURES = frozenset({"injection_probe.docx"})
_DOC_IDS = [doc.doc_id for doc in GOLDEN_SET]
_LOCATED_DOCS = [doc for doc in GOLDEN_SET if doc.locators]
_WORKSTREAM_DOC_PAIRS: list[tuple[Workstream, str]] = sorted(
    DEEP_WORKSTREAM_DOCUMENTS.items(), key=lambda pair: pair[1]
)


def _committed_corpus_ids() -> set[str]:
    ids: set[str] = set()
    for base in (DATA_DIR, SCENARIOS_DIR):
        for path in base.iterdir():
            if path.suffix.lower() in _PARSEABLE_SUFFIXES and path.name not in _REDTEAM_FIXTURES:
                ids.add(path.name)
    return ids


class TestGoldenSetShape:
    def test_golden_set_has_twenty_unique_docs(self) -> None:
        assert len(GOLDEN_SET) == 20
        doc_ids = [doc.doc_id for doc in GOLDEN_SET]
        assert len(set(doc_ids)) == 20

    def test_golden_set_exactly_covers_committed_clean_corpus(self) -> None:
        assert {doc.doc_id for doc in GOLDEN_SET} == _committed_corpus_ids()

    def test_exactly_four_keystones_carry_expectations(self) -> None:
        with_findings = [doc for doc in GOLDEN_SET if doc.expected_finding_titles]
        assert {doc.doc_id for doc in with_findings} == set(DEEP_WORKSTREAM_DOCUMENTS.values())
        bare = [doc for doc in GOLDEN_SET if not doc.expected_finding_titles]
        assert len(bare) == 16
        assert all(doc.expected_entities == () for doc in bare)

    def test_injection_probe_is_not_pinned(self) -> None:
        assert "injection_probe.docx" not in {doc.doc_id for doc in GOLDEN_SET}

    def test_loader_helpers(self) -> None:
        assert load_golden_set() is GOLDEN_SET
        assert golden_doc("contract_meridian_logistics.pdf").locators == (COC_CLAUSE_LOCATOR,)
        with pytest.raises(KeyError):
            golden_doc("not_a_pinned_doc.pdf")
        with pytest.raises(FileNotFoundError):
            golden_path("not_a_pinned_doc.pdf")


class TestCorpusOnDisk:
    @pytest.mark.parametrize("golden", GOLDEN_SET, ids=_DOC_IDS)
    def test_doc_exists_and_parses_or_flagged_needs_ocr(self, golden: GoldenDoc) -> None:
        path = golden_path(golden.doc_id)
        parsed = LocalParser().parse(path.read_bytes(), golden.doc_id, DEAL)
        assert parsed.format.needs_ocr is golden.needs_ocr
        if golden.needs_ocr:
            assert parsed.text is None
            assert chunk(parsed) == ()
        else:
            assert parsed.text is not None

    @pytest.mark.parametrize("golden", GOLDEN_SET, ids=_DOC_IDS)
    def test_parsed_docs_contain_no_lorem(self, golden: GoldenDoc) -> None:
        parsed = LocalParser().parse(golden_path(golden.doc_id).read_bytes(), golden.doc_id, DEAL)
        if parsed.text is not None:
            assert "lorem" not in parsed.text.lower(), f"{golden.doc_id} contains lorem filler"


class TestKeystoneLocatorPins:
    @pytest.mark.parametrize("golden", _LOCATED_DOCS, ids=[doc.doc_id for doc in _LOCATED_DOCS])
    def test_pinned_locators_resolve(self, golden: GoldenDoc) -> None:
        parsed = LocalParser().parse(golden_path(golden.doc_id).read_bytes(), golden.doc_id, DEAL)
        resolved = {c.locator for c in chunk(parsed)}
        for locator in golden.locators:
            assert locator in resolved, f"{golden.doc_id}: locator {locator!r} does not resolve"

    def test_coc_span_extractable_at_clause_11_3(self) -> None:
        parsed = LocalParser().parse(
            golden_path("contract_meridian_logistics.pdf").read_bytes(),
            "contract_meridian_logistics.pdf",
            DEAL,
        )
        assert parsed.text is not None
        target = next(c for c in chunk(parsed) if c.locator == COC_CLAUSE_LOCATOR)
        assert COC_SPAN in " ".join(target.text.split())
        assert COC_SPAN in " ".join(parsed.text.split())

    def test_concentration_ratio_pinned_at_18_3(self) -> None:
        assert MERIDIAN_REVENUE / TOTAL_REVENUE == GOLDEN_CONCENTRATION_RATIO
        assert f"{GOLDEN_CONCENTRATION_RATIO:.3%}" == "18.300%"
        workbook = load_workbook(golden_path("financials_fy27.xlsx"), data_only=True)
        assert FY27_REVENUE_SHEET in workbook.sheetnames
        sheet = workbook[FY27_REVENUE_SHEET]
        meridian_rows = [
            row for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None
        ]
        meridian = next(row for row in meridian_rows if row[0] == "Meridian Logistics, Inc.")
        total = next(row for row in meridian_rows if row[0] == "TOTAL")
        assert isinstance(meridian[2], int | float)
        assert isinstance(total[2], int | float)
        assert meridian[2] / total[2] == GOLDEN_CONCENTRATION_RATIO

    def test_whitfield_departure_is_roster_date_plus_60d(self) -> None:
        assert (WHITFIELD_DEPARTURE - ROSTER_REFERENCE_DATE).days == 60
        workbook = load_workbook(golden_path("hr_roster_vantage.xlsx"))
        sheet = workbook["Roster"]
        whitfield = next(
            row
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[0] == "Dana Whitfield"
        )
        departure = whitfield[5]
        assert isinstance(departure, datetime)
        assert departure.date() == WHITFIELD_DEPARTURE

    def test_titanbridge_eol_locator_carries_the_fact(self) -> None:
        golden = golden_doc("tech_inventory.pdf")
        parsed = LocalParser().parse(golden_path(golden.doc_id).read_bytes(), golden.doc_id, DEAL)
        entry = next(c for c in chunk(parsed) if c.locator in golden.locators)
        assert "TitanBridge 4.1" in entry.text
        assert "end-of-life" in entry.text


class TestKeystoneProducers:
    @pytest.mark.parametrize(
        ("workstream", "doc_id"),
        _WORKSTREAM_DOC_PAIRS,
        ids=[doc_id for _, doc_id in _WORKSTREAM_DOC_PAIRS],
    )
    def test_producer_title_and_entities_match_golden_pin(
        self, workstream: Workstream, doc_id: str
    ) -> None:
        golden = golden_doc(doc_id)
        parsed = LocalParser().parse(golden_path(doc_id).read_bytes(), doc_id, DEAL)
        fact = extract_fact(workstream, parsed)
        assert fact.title in golden.expected_finding_titles
        assert fact.affected_entities == golden.expected_entities
