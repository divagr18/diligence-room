"""Cross-workstream query tool (BUILD_PLAN D5-M2, vision §6/§7.5).

Agents never talk directly: ``ask_agent`` submits a question through the
gateway, which decides policy, calls the target workstream's responder, and
filters the response down to its allowed shape. ``OfflineFinanceResponder``
computes the real Customer X revenue share from the dataset workbook — no
hardcoded answers in the demo path.
"""

from __future__ import annotations

import json
import uuid
from collections.abc import Mapping
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Protocol

from google.cloud import firestore
from openpyxl import load_workbook

from gateway.aggregate import (
    AggregateAnswer,
    ExtractionBlocked,
    enforce_response_shape,
    render_aggregate,
    screen_question,
)
from gateway.decide import DecisionReason, GatewayRequest, Verdict, decide
from gateway.policy import PolicyStore
from identity.principals import Principal
from memory.event_log import EventLog
from registry.models import Workstream
from runtime.events import EventType, new_event

_DEFAULT_WORKBOOK = (
    Path(__file__).resolve().parents[2] / "data" / "acme_robotics" / "financials_fy27.xlsx"
)
_FINANCE_SHEET = "FY27 Projected Revenue"
_CUSTOMER_X_ALIAS = "Meridian"


@dataclass(frozen=True, slots=True)
class GatewayResponse:
    """Outcome of one governed query."""

    request_id: str
    verdict: Verdict
    reason: DecisionReason
    answer: str | None
    responder: str

    def to_json(self) -> str:
        return json.dumps(
            {
                "request_id": self.request_id,
                "decision": self.verdict.value,
                "reason": self.reason.value,
                "answer": self.answer,
                "responder": self.responder,
            },
            sort_keys=True,
        )


class TargetResponder(Protocol):
    """A workstream endpoint that answers policy-approved questions."""

    def answer(self, deal_id: str, question: str, purpose: str) -> str: ...


class AskAgentTool(Protocol):
    """ADK-callable tool surface with named parameters."""

    __name__: str

    def __call__(self, target_ws: str, question: str, purpose: str) -> dict[str, str]: ...


class OfflineFinanceResponder:
    """Deterministic finance responder backed by the dataset workbook.

    Computes Customer X's share of projected FY27 revenue (Meridian / TOTAL)
    on every call — the 18.3% headline is always derived, never stored.
    """

    def __init__(self, workbook_path: Path = _DEFAULT_WORKBOOK) -> None:
        self._workbook_path = workbook_path

    def compute_share(self) -> AggregateAnswer:
        workbook = load_workbook(self._workbook_path, data_only=True)
        sheet = workbook[_FINANCE_SHEET]
        customer_x_revenue: float | None = None
        total_revenue: float | None = None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            customer = str(row[0]) if row[0] is not None else ""
            revenue = row[2]
            if not isinstance(revenue, int | float):
                continue
            if customer == "TOTAL":
                total_revenue = float(revenue)
            elif _CUSTOMER_X_ALIAS in customer:
                customer_x_revenue = float(revenue)
        if customer_x_revenue is None or total_revenue in (None, 0.0):
            raise ValueError(f"cannot compute revenue share from {self._workbook_path}")
        share_percent = customer_x_revenue / total_revenue * 100.0
        return AggregateAnswer(
            metric="customer_x_revenue_share",
            value=share_percent,
            unit="percent",
            source_document=self._workbook_path.name,
            basis=f"{_FINANCE_SHEET} sheet (Customer X / TOTAL)",
        )

    def answer(self, deal_id: str, question: str, purpose: str) -> str:
        del deal_id, question, purpose
        return render_aggregate(self.compute_share())


@dataclass(frozen=True, slots=True)
class LocalGatewayClient:
    """In-process gateway: decide -> respond -> filter, with full audit."""

    client: firestore.Client
    responders: Mapping[Workstream, TargetResponder] = field(default_factory=dict)

    def ask(
        self,
        sender: Principal,
        deal_id: str,
        target: Workstream,
        question: str,
        purpose: str,
        ts: datetime | None = None,
    ) -> GatewayResponse:
        stamp = ts if ts is not None else datetime.now(UTC)
        request_id = uuid.uuid4().hex
        try:
            screen_question(question)
        except ExtractionBlocked as blocked:
            self._audit_block(sender, deal_id, target, purpose, request_id, stamp)
            return GatewayResponse(
                request_id=request_id,
                verdict=Verdict.DENY,
                reason=blocked.reason,
                answer=None,
                responder="-",
            )
        request = GatewayRequest(
            request_id=request_id,
            deal_id=deal_id,
            sender=sender,
            target_workstream=target,
            question=question,
            purpose=purpose,
            ts=stamp,
        )
        decision = decide(self.client, request, now=stamp)
        if decision.verdict is Verdict.DENY:
            return GatewayResponse(
                request_id=request_id,
                verdict=decision.verdict,
                reason=decision.reason,
                answer=None,
                responder="-",
            )
        responder = self.responders.get(target)
        if responder is None:
            raise LookupError(f"no responder registered for workstream {target.value!r}")
        raw = responder.answer(deal_id, question, purpose)
        shape = PolicyStore(self.client).get(deal_id, sender.workstream, target)
        if shape is None:
            raise LookupError(f"policy rule vanished after ALLOW for {request_id}")
        try:
            filtered = enforce_response_shape(raw, shape.response_shape)
        except ExtractionBlocked as blocked:
            self._audit_block(sender, deal_id, target, purpose, request_id, stamp)
            return GatewayResponse(
                request_id=request_id,
                verdict=Verdict.DENY,
                reason=blocked.reason,
                answer=None,
                responder=target.value,
            )
        return GatewayResponse(
            request_id=request_id,
            verdict=Verdict.ALLOW,
            reason=DecisionReason.AGGREGATE_PERMITTED,
            answer=filtered,
            responder=target.value,
        )

    def _audit_block(
        self,
        sender: Principal,
        deal_id: str,
        target: Workstream,
        purpose: str,
        request_id: str,
        now: datetime,
    ) -> None:
        event = new_event(
            deal_id=deal_id,
            actor=sender.name,
            event_type=EventType.GATEWAY_DECISION,
            payload={
                "decision": Verdict.DENY.value,
                "reason": DecisionReason.RAW_MODEL_PROHIBITED.value,
                "subject": sender.name,
                "target": target.value,
                "purpose": purpose,
                "request_id": request_id,
            },
            now=now,
        )
        EventLog(self.client).append(event)


def gateway_query_tool(
    principal: Principal, deal_id: str, client: LocalGatewayClient
) -> AskAgentTool:
    """Build the ADK ``ask_agent`` tool bound to one principal and deal."""

    def ask_agent(target_ws: str, question: str, purpose: str) -> dict[str, str]:
        """Ask another workstream a policy-governed question via the gateway.

        Args:
            target_ws: Target workstream id (e.g. "finance").
            question: The question to ask the target workstream.
            purpose: Declared purpose; must match a policy allow-list entry.

        Returns:
            Dict with "decision" (allow/deny), machine-readable "reason", and
            "answer" (empty string when denied).
        """
        try:
            target = Workstream(target_ws)
        except ValueError:
            raise ValueError(f"unknown workstream {target_ws!r}") from None
        response = client.ask(
            sender=principal,
            deal_id=deal_id,
            target=target,
            question=question,
            purpose=purpose,
        )
        return {
            "decision": response.verdict.value,
            "reason": response.reason.value,
            "answer": response.answer or "",
        }

    return ask_agent
