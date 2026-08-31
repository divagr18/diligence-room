"""Document serving + evidence locator for the Deal Room document viewer.

Serves data-room documents and locates the exact page (PDF) or sheet+row
(XLSX) containing an evidence ``verbatim_span``, so the frontend can open the
source at the precise place the offending line lives. Text matching reuses the
ingestion parser's extraction so a span already verified against the parsed
text is guaranteed to resolve.
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from ingestion.classifier import FakeClassifier
from ingestion.formats import detect_format
from ingestion.models import FormatKind

_REPO_ROOT = Path(__file__).resolve().parents[2]
DATA_ROOM = _REPO_ROOT / "data" / "vantage_robotics"

_ws = re.compile(r"\s+")


def _norm(text: str) -> str:
    return _ws.sub(" ", text).strip()


def resolve_document_path(document_id: str) -> Path | None:
    """Resolve a data-room document to a real path, or None if absent/unsafe.

    Names the OS refuses to resolve (NUL bytes, illegal characters) fail
    closed to None — the route answers 404, never 500.
    """
    try:
        candidate = (DATA_ROOM / document_id).resolve()
    except (OSError, ValueError):
        return None
    if DATA_ROOM.resolve() not in candidate.parents and candidate.parent != DATA_ROOM.resolve():
        return None
    return candidate if candidate.is_file() else None


def _pdf_text_pages(blob: bytes) -> list[str]:
    from io import BytesIO

    from pypdf import PdfReader

    reader = PdfReader(BytesIO(blob))
    return [(page.extract_text() or "") for page in reader.pages]


def _locate_pdf(blob: bytes, span: str) -> dict[str, object]:
    pages = _pdf_text_pages(blob)
    needle = _norm(span)
    page: int | None = None
    for index, text in enumerate(pages):
        if needle and needle in _norm(text):
            page = index + 1
            break
    return {"kind": "pdf", "page": page, "page_count": len(pages)}


def _xlsx_rows(blob: bytes) -> list[tuple[str, tuple[tuple[str, ...], ...]]]:
    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(blob), data_only=True)
    sheets: list[tuple[str, tuple[tuple[str, ...], ...]]] = []
    for sheet in workbook.worksheets:
        rows = tuple(
            tuple("" if cell is None else str(cell) for cell in row)
            for row in sheet.iter_rows(values_only=True)
            if not all(cell is None for cell in row)
        )
        sheets.append((sheet.title, rows))
    return sheets


def _locate_xlsx(blob: bytes, span: str) -> dict[str, object]:
    needle = _norm(span)
    for sheet_name, rows in _xlsx_rows(blob):
        for row_index, row in enumerate(rows):
            if needle and needle in _norm("\t".join(row)):
                headers = list(rows[0]) if rows else []
                return {
                    "kind": "xlsx",
                    "sheet": sheet_name,
                    "row_index": row_index,
                    "headers": headers,
                    "rows": [list(row) for row in rows],
                }
    return {"kind": "xlsx", "sheet": None, "row_index": None, "headers": [], "rows": []}


def locate_evidence(document_id: str, span: str) -> dict[str, object] | None:
    """Locate *span* within *document_id*. Returns locator dict or None."""
    path = resolve_document_path(document_id)
    if path is None:
        return None
    blob = path.read_bytes()
    info = detect_format(blob, document_id)
    if info.kind is FormatKind.NATIVE_PDF or info.kind is FormatKind.SCANNED_PDF:
        return _locate_pdf(blob, span)
    if info.kind is FormatKind.XLSX:
        return _locate_xlsx(blob, span)
    return {"kind": info.kind.value, "page": None, "page_count": None}


# --------------------------------------------------------------------------
# Data-room listing
# --------------------------------------------------------------------------

# Everything in the data room is a document except the plan that describes it.
_NOT_A_DOCUMENT = {"DATASET_PLAN.md"}


@dataclass(frozen=True, slots=True)
class DocumentSummary:
    """One data-room file, as the Documents view needs it."""

    document_id: str
    format: str
    mime: str
    needs_ocr: bool
    size_bytes: int
    page_count: int | None
    checksum: str
    doc_type: str
    workstream: str | None
    confidence: float


def _extract_text(blob: bytes, kind: FormatKind) -> str:
    """Best-effort text for routing. A parser failure costs a label, not the row."""
    try:
        if kind in (FormatKind.NATIVE_PDF, FormatKind.SCANNED_PDF):
            return "\n".join(_pdf_text_pages(blob))
        if kind is FormatKind.XLSX:
            return "\n".join("\t".join(row) for _sheet, rows in _xlsx_rows(blob) for row in rows)
    except Exception:  # noqa: BLE001 - an unreadable file still belongs in the list
        return ""
    return ""


def _page_count(blob: bytes, kind: FormatKind) -> int | None:
    try:
        if kind in (FormatKind.NATIVE_PDF, FormatKind.SCANNED_PDF):
            return len(_pdf_text_pages(blob))
        if kind is FormatKind.XLSX:
            return len(_xlsx_rows(blob))
    except Exception:  # noqa: BLE001
        return None
    return None


@lru_cache(maxsize=1)
def list_documents() -> tuple[DocumentSummary, ...]:
    """Every file in the data room, with format, routing and page count.

    Cached: the data room is static per deployment, and building the list parses
    every PDF and workbook. The routing label comes from the same deterministic
    classifier the ingestion pipeline uses, so the workstream shown here is the
    one a document would actually be routed to rather than a second opinion.
    """
    if not DATA_ROOM.is_dir():
        return ()
    classifier = FakeClassifier()
    summaries: list[DocumentSummary] = []
    for path in sorted(DATA_ROOM.iterdir()):
        if not path.is_file() or path.name.startswith(".") or path.name in _NOT_A_DOCUMENT:
            continue
        blob = path.read_bytes()
        info = detect_format(blob, path.name)
        decision = classifier.classify(path.name, _extract_text(blob, info.kind), None)
        summaries.append(
            DocumentSummary(
                document_id=path.name,
                format=info.kind.value,
                mime=info.mime,
                needs_ocr=info.needs_ocr,
                size_bytes=len(blob),
                page_count=_page_count(blob, info.kind),
                checksum=hashlib.sha256(blob).hexdigest(),
                doc_type=decision.doc_type,
                workstream=decision.workstream,
                confidence=decision.confidence,
            )
        )
    return tuple(summaries)
